# -*- coding: utf-8 -*-

import openpyxl

__author__ = 'wangyongjun'


class Excel:
    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.load_workbook(self.path)
        # self.ws = self.wb['Sheet0']
        self.ws = self.wb.active
        # self.ws.protection.sheet = False

    def read_by_cell(self, rowx, colx):
        return self.ws.cell(row=rowx, column=colx).value

    def read_by_col(self, col):
        return self.ws[col]

    def read_by_row(self, row):
        return self.ws[row]

    def write_data(self, w_dict, specific_path=None):
        for loc in w_dict:
            self.ws[loc] = w_dict[loc]

        if specific_path:
            self.wb.save(specific_path)
        else:
            self.wb.save(self.path)


if __name__ == "__main__":
    pass
